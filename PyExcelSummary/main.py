
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


RED = 'ff4d4d'
GREEN = '66ff66'
GRAY = 'bfbfbf'
BLACK = '000001'
WHITE = 'ffffff'
ORANGE = 'ffa500'


def excelReport():
    print("Please provide the path to the file:")
    path = input()
    path = path.replace('"', '')
    wb = load_workbook(path)
    i = len(wb.sheetnames)

    for sheet in wb.worksheets:
        total_failed = 0
        total_passed = 0
        total_not_performed = 0
        total_test_cases = 0
        total_passed_with_deviation = 0

        wb = load_workbook(path)
        currentSheetName = wb.sheetnames[len(wb.sheetnames) - i]
        ws = wb[wb.sheetnames[len(wb.sheetnames) - i]]

        if currentSheetName == "DocHistory" or currentSheetName == "Test_Summary" or currentSheetName == "Status":
            i += 1
            continue
        else:
            for rows in ws.iter_rows():
                for cell in rows:
                    if str(cell.value).lower() == 'failed':
                        cell.fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
                        cell.font = Font(bold=True, color=GREEN, size=16)
                        total_failed += 1
                        total_test_cases += 1
                    if str(cell.value).lower() == 'passed':
                        if str(ws.cell(row=cell.row, column=cell.column + 1).value).lower() != 'none':
                            total_passed_with_deviation += 1
                            total_passed -= 1
                        cell.fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
                        cell.font = Font(bold=True, color=RED, size=16)
                        total_passed += 1
                        total_test_cases += 1
                    if str(cell.value).lower() == 'not performed':
                        cell.fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type="solid")
                        cell.font = Font(bold=True, color=WHITE, size=16)
                        total_not_performed += 1
                        total_test_cases += 1
        print(
            f"{currentSheetName} \n 'passed:{total_passed} (minor deviation: {total_passed_with_deviation})' 'failed:{total_failed}' 'not_performed:{total_not_performed}' \n")

        ws = wb['Test_Summary']
        for rows in ws.iter_rows():
            for cell in rows:
                currentSheetName = wb.sheetnames[len(wb.sheetnames) - i]
                if str(cell.value) == str(currentSheetName):
                    ws.cell(row=cell.row, column=cell.column + 1).value = total_passed
                    ws.cell(row=cell.row, column=cell.column + 2).value = total_passed_with_deviation
                    ws.cell(row=cell.row, column=cell.column + 3).value = total_failed
                    ws.cell(row=cell.row, column=cell.column + 4).value = total_not_performed
                    ws.cell(row=cell.row, column=cell.column + 5).value = total_test_cases
        i += 1
        wb.save(path)

    print(f"Done.")


excelReport()
